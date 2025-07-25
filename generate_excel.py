import openpyxl
import ast

# Create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Validation Results"

# Set headers
ws['A1'] = 'FAILED TESTS'
ws['B1'] = 'CUSTOM TEMPLATE PARAMETERS'
ws['C1'] = 'MICROSOFT TEMPLATE PARAMETERS'
ws['D1'] = 'SUGGESTED PARAMETERS'

# Read failed tests and suggested parameters from results.txt
failed_tests = []
custom_params = []
qs_params = []
suggested_params= []

try:
    with open("results.txt", "r", encoding="utf-8") as file:
        lines = file.readlines()
        for line in lines:
            if "[-]" in line:
                failed_tests.append(line.strip())
            elif line.startswith("Custom params:"):
                param_str = line.split("Custom params:")[1].strip()
                try:
                    custom_params = list(ast.literal_eval(param_str))
                except Exception as e:
                    print(f"Error parsing custom params: {e}")
            elif line.startswith("QS params:"):
                param_str = line.split("QS params:")[1].strip()
                try:
                    qs_params = list(ast.literal_eval(param_str))
                except Exception as e:
                    print(f"Error parsing QS params: {e}")
            elif line.startswith ("Suggested params:"):
                param_str = param_str = line.split("Suggested params:")[1].strip()
                try:
                    suggested_params = list(ast.literal_eval(param_str))
                except Exception as e:
                    print(f"Error parsing suggested params: {e}")

except FileNotFoundError:
    print("results.txt file not found.")

# Determine the maximum number of rows needed
max_rows = max(len(failed_tests), len(custom_params), len(qs_params))

# Write data to Excel
for i in range(max_rows):
    if i < len(failed_tests):
        ws.cell(row=i+2, column=1, value=failed_tests[i])
    if i < len(custom_params):
        ws.cell(row=i+2, column=2, value=custom_params[i])
    if i < len(qs_params):
        ws.cell(row=i+2, column=3, value=qs_params[i])
    if i < len(suggested_params):
        ws.cell(row=i+2, column=4, value=suggested_params[i])

# Save the workbook
wb.save("validation_results.xlsx")
print("Excel file 'validation_results.xlsx' has been created successfully.")

